# -*- coding: utf-8 -*-
"""
Created on Fri May 21 15:51:28 2021

@author: leokt
"""
# =============================================================================
# Steps:

# pre-analyzed:
# mouse IMCD RNA seq (Chen et al 2021)
# IMCD suspension proteome (this paper)
# IMCD microdissected proteome (Limbutara et al 2020)
#
# Screening for S/T kinases (Sugiyama et al 2019)
# Colocalization from fractions (Yang et al 2015)
# Known kinases
# Position ranking (Sugiyama et al 2019) only for selected positions
# =============================================================================

import pandas as pd
import numpy as np
from bayesmodules import interpolate
from bayesmodules import cMBF_calc


fileloc_1 = "/Users/kirbyleo/Box Sync/Depot - dDAVP-time course - Kirby/BayesAnalysis/expression data.xlsx"
express_df = pd.read_excel(fileloc_1, None)

#from sugiyama_data import STYdotscores_df
main_vec = express_df['exp bayes calc V2']
STYdots = interpolate(STYdotscores_df)
main_vec = main_vec.rename(columns = {'Gene Symbol':'Kinases'})

# %%
#STY ranking
main_vec = main_vec.merge(STYdots,on='Kinases',how='left')
clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
for i in clus_names: 
    STY_cMBF = cMBF_calc(main_vec[i],1/9)
    prod = STY_cMBF*main_vec['Posterior_Exp']
    main_vec[i + '_STY'] = prod/sum(prod)
main_vec = main_vec.drop(clus_names,axis=1)

fileloc_3 = "/Users/kirbyleo/Box Sync/Depot - dDAVP-time course - Kirby/BayesAnalysis/BayesTC complete calc_python V17.xlsx"
writer=pd.ExcelWriter(fileloc_3)
main_vec.to_excel(writer, sheet_name='STY ranking')

writer.save()

# %%
#colocalization ranking
fileloc_2 = "/Users/kirbyleo/Box Sync/Depot - dDAVP-time course - Kirby/BayesAnalysis/colocalization mapping V2.xlsx"
coloc_df = pd.read_excel(fileloc_2, "coloc dot ranking V2")

coloc_vec = main_vec[['Kinases']].copy()
coloc_vec = coloc_vec.merge(coloc_df,on='Kinases',how='left')
for i in clus_names:
    coloc_cMBF = cMBF_calc(coloc_vec[i],np.mean(coloc_vec[i]))
    prod1 = coloc_cMBF*main_vec[i + '_STY']
    coloc_vec[i + '_coloc'] = prod1/sum(prod1)
coloc_vec = coloc_vec.drop(clus_names,axis=1)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
coloc_vec.to_excel(writer, sheet_name = 'coloc ranking')
writer.save()
writer.close()

# %%
# known kinases
fileloc_4 = "/Users/kirbyleo/Box Sync/Depot - dDAVP-time course - Kirby/BayesAnalysis/Bayes sheets/BayesTC_known kinases.xlsx"
known_df = pd.read_excel(fileloc_4, "Simplified combined V3")

cluster_direct = ['Increase','Decrease','Decrease','Decrease','Decrease','Increase','Increase','Decrease',
                  'Decrease','Increase','Decrease','Decrease','Increase','Decrease']

clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
position_based = coloc_vec
known_based = coloc_vec[['Kinases']].copy()
for a in range(len(clus_names)):
    new_likelihood = list(range(len(position_based['Kinases'])))
    for b in range(len(position_based['Kinases'])):
        if position_based['Kinases'][b] in known_df['Kinases'].values:
            if known_df[known_df['Kinases']==position_based['Kinases'][b]]['Net Effect on Activity'].tolist()[0] == cluster_direct[a]:
                new_likelihood[b] = 0.9
            elif known_df[known_df['Kinases']==position_based['Kinases'][b]]['Net Effect on Activity'].tolist()[0] == 'Regulated':
                new_likelihood[b] = 0.7 
            else:
                new_likelihood[b] = 0.1
        else:
            new_likelihood[b] = 0.5
    known_based[clus_names[a]] = new_likelihood

known_vec = main_vec[['Kinases']].copy()
for c in clus_names:
    prod2 = position_based[c + "_coloc"]*known_based[c]
    new_prior = prod2/sum(prod2)
    known_vec[c+'_known'] = new_prior.tolist()
    
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
known_vec.to_excel(writer, sheet_name = 'known kinase ranking')
writer.save()
writer.close()

# %%
#position ranking
#from sugiyama_data import pos_dotscore_list
posit = ['-6','-5','-4','-3','-2','-1','+1','+2','+3','+4','+5','+6']

position_based = main_vec[['Kinases']].copy()
int_pos = list(range(len(pos_dotscore_list)))
for t in range(len(pos_dotscore_list)):
    int_pos[t] = interpolate(pos_dotscore_list[t])

from bayesmodules import selectpos_bayes
from bayesmodules import cMBF_calc
bayes_IB = selectpos_bayes('IB',int_pos,known_vec)

bayes_output = bayes_IB[['Kinases']].copy()  
bayes_output['IB'] = bayes_IB.iloc[:,-1:]

book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
bayes_IB.to_excel(writer, sheet_name = 'IB')
writer.save()
writer.close()

clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
rest_clus = clus_names
rest_clus.remove('IB')
from openpyxl import load_workbook
for v in rest_clus:
    bayes_clus = selectpos_bayes(v,int_pos,known_vec)
    bayes_output[v] = bayes_clus.iloc[:,-1:]
    book = load_workbook(fileloc_3)
    writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
    writer.book = book
    bayes_clus.to_excel(writer, sheet_name = v)
    writer.save()
    writer.close()

book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
bayes_output.to_excel(writer, sheet_name = 'all clusters')
writer.save()
writer.close()



### check this later
if __name__ == "__main__":
   # stuff only to run when not called via 'import' here
    try:
        main()
    except SystemExit:
        raise
    except:
        print_exc()
        sys.exit(-1)
